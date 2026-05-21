# services/preencher_codigo_beline.py
from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Callable, Optional

import openpyxl


# Ex.: "104-. DS Beline - JUDICIAL - Anderson Cesar Rodrigues X INSS - MA - Proc"
# Aceita NUM_ATUAL vazio (104-.)
PREFIX_RE = re.compile(
    r"^\s*(?P<base>\d+)\s*-\s*(?P<atual>\d*)\s*\.\s*(?P<resto>.+?)\s*$",
    flags=re.IGNORECASE,
)

# Extrai o "cliente" após "- JUDICIAL -" e antes de " x " (ou " X ")
CLIENTE_RE = re.compile(
    r"\s-\sJUDICIAL\s-\s(?P<cliente>.+?)\s+[xX]\s+",
    flags=re.IGNORECASE,
)


def _norm_key(s: str) -> str:
    """
    Normalização forte para bater nomes:
    - remove acentos
    - UPPERCASE
    - remove pontuação
    - colapsa espaços
    """
    s = (s or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _extract_cliente_from_folder_name(folder_name: str) -> Optional[str]:
    """
    Regra principal:
      cliente = parte depois de " - JUDICIAL - " e antes de " x " / " X "
    """
    m = CLIENTE_RE.search(folder_name)
    if not m:
        return None
    cliente = m.group("cliente").strip()
    return cliente or None


def _extract_uf_from_folder_name(folder_name: str) -> Optional[str]:
    """
    Tenta extrair UF no padrão "... - MA - ..." (dois caracteres).
    Retorna None se não achar.
    """
    m = re.search(r"\s-\s([A-Z]{2})\s-\s", folder_name)
    return m.group(1).strip().upper() if m else None


def _find_header_map(ws) -> dict[str, int]:
    """
    Mapeia headers da linha 1 (normalizados) -> índice (1-based)
    """
    col_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        col_map[_norm_key(str(v))] = c
    return col_map


def _coerce_int_like(value: object) -> str:
    """
    Converte valores numéricos do Excel para string sem .0 quando aplicável.
    """
    if value is None:
        return ""
    s = str(value).strip()
    # Ex.: "26459.0" -> "26459"
    if re.fullmatch(r"\d+\.0", s):
        return s[:-2]
    return s


def _build_excel_lookup(
    excel_path: Path,
    sheet_name: str,
    col_nome: str,
    col_numero: str,
    col_uf: str = "ESTADO",
) -> tuple[dict[str, list[tuple[str, Optional[str]]]], list[str]]:
    """
    Retorna:
      lookup: key(nome_normalizado) -> lista[(numero_novo, uf_ou_None)]
      avisos: mensagens para log
    """
    avisos: list[str] = []

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}, [f"Excel: aba '{sheet_name}' não encontrada. Abas: {wb.sheetnames}"]

    ws = wb[sheet_name]
    headers = _find_header_map(ws)

    c_nome = headers.get(_norm_key(col_nome))
    c_num = headers.get(_norm_key(col_numero))
    c_uf = headers.get(_norm_key(col_uf))  # opcional

    if c_nome is None:
        return {}, [f"Excel: não achei a coluna '{col_nome}' (na aba '{sheet_name}')."]
    if c_num is None:
        return {}, [f"Excel: não achei a coluna '{col_numero}' (na aba '{sheet_name}')."]

    lookup: dict[str, list[tuple[str, Optional[str]]]] = {}

    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, c_nome).value
        num = ws.cell(r, c_num).value

        nome_s = ("" if nome is None else str(nome)).strip()
        num_s = _coerce_int_like(num)

        if not nome_s:
            continue
        if not num_s:
            avisos.append(f"Excel: linha {r} sem '{col_numero}' para '{nome_s}'.")
            continue

        uf_s: Optional[str] = None
        if c_uf is not None:
            uf_val = ws.cell(r, c_uf).value
            if uf_val is not None:
                uf_s = str(uf_val).strip().upper() or None

        lookup.setdefault(_norm_key(nome_s), []).append((num_s, uf_s))

    return lookup, avisos


def build_plan_codigo_beline(
    pasta_raiz: Path,
    excel_path: Path,
    sheet_name: str = "Dados",
    col_nome: str = "CLIENTE",
    col_numero: str = "Nº PARCEIRO",
    log: Optional[Callable[[str], None]] = None,
) -> list[tuple[Path, str]]:
    """
    Monta um plan de renomeação (compatível com apply_rename_plan):
      plan: List[(Path_da_pasta, novo_nome_da_pasta)]

    Regras:
    - Pasta no formato: NUM_BASE-NUM_ATUAL. RESTO
      (aceita NUM_ATUAL vazio: NUM_BASE-. RESTO)
    - Extrai cliente a partir do nome: após "- JUDICIAL -" e antes de " x " / " X "
    - Dá match com Excel (col_nome)
    - Substitui SOMENTE o número após hífen (NUM_ATUAL) pelo número do Excel (col_numero)
    - Mantém todo o resto idêntico.
    - Se houver mais de 1 candidato no Excel para o mesmo nome:
        tenta desempatar por UF (quando existir no nome da pasta e no Excel via coluna ESTADO)
        se não conseguir, marca como AMBIGUO e não renomeia.
    """
    def _log(msg: str) -> None:
        if log:
            log(msg)

    if not pasta_raiz.exists():
        _log(f"ERRO: pasta base não existe: {pasta_raiz}")
        return []
    if not excel_path.exists():
        _log(f"ERRO: Excel não existe: {excel_path}")
        return []

    lookup, avisos = _build_excel_lookup(
        excel_path=excel_path,
        sheet_name=sheet_name,
        col_nome=col_nome,
        col_numero=col_numero,
    )
    for a in avisos:
        _log(a)

    if not lookup:
        _log("ERRO: lookup do Excel ficou vazio (sem dados válidos).")
        return []

    plan: list[tuple[Path, str]] = []

    pastas = [p for p in pasta_raiz.iterdir() if p.is_dir()]
    _log(f"Pastas encontradas: {len(pastas)}")

    for p in sorted(pastas, key=lambda x: x.name.lower()):
        nome_atual = p.name

        m = PREFIX_RE.match(nome_atual)
        if not m:
            _log(f"PULAR: '{nome_atual}' (não bate padrão NUM_BASE-NUM_ATUAL.)")
            continue

        num_base = m.group("base")
        num_atual = (m.group("atual") or "").strip()
        resto = m.group("resto")

        cliente = _extract_cliente_from_folder_name(nome_atual)
        if not cliente:
            _log(f"PULAR: '{nome_atual}' (não consegui extrair cliente pelo padrão JUDICIAL)")
            continue

        key = _norm_key(cliente)
        candidatos = lookup.get(key)
        if not candidatos:
            _log(f"SEM MATCH: '{nome_atual}' (cliente='{cliente}')")
            continue

        # Decide número novo
        numero_novo: Optional[str] = None
        if len(candidatos) == 1:
            numero_novo = candidatos[0][0]
        else:
            # tenta por UF
            uf_pasta = _extract_uf_from_folder_name(nome_atual)
            if uf_pasta:
                filtrados = [num for (num, uf) in candidatos if (uf or "").upper() == uf_pasta]
                if len(filtrados) == 1:
                    numero_novo = filtrados[0]

        if not numero_novo:
            _log(
                f"AMBIGUO: '{nome_atual}' (cliente='{cliente}', candidatos={candidatos})"
            )
            continue

        # Se já está igual (e num_atual não é vazio), não precisa renomear
        if num_atual and numero_novo.strip() == num_atual.strip():
            _log(f"OK (já está certo): {nome_atual}")
            continue

        novo_nome = f"{num_base}-{numero_novo}. {resto}"
        plan.append((p, novo_nome))
        _log(f"CASADA: {nome_atual} -> {novo_nome}")

    _log(f"Resumo plan: casadas={len(plan)}, sem_match/ignoradas veja acima")
    return plan
