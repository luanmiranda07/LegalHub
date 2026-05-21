from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class Registro:
    numero_parceiro: str = ""
    parceiro: str = ""
    cliente: str = ""
    cpf: str = ""
    reu: str = ""
    estado: str = ""

    # NOVO: vem da planilha base (merge)
    id_base: str = ""

    # NÃO preencher automaticamente
    classificacao: str = ""
    observacao: str = ""
    responsavel: str = ""

    nome_pasta: str = ""

    # Para validação/debug
    partes_brutas: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


UF_RE = re.compile(r"^[A-Z]{2}$")

# Separador "de verdade" entre campos: hífen com pelo menos 1 espaço antes e depois
DASH_SEP_RE = re.compile(r"\s+-\s+")

# ID (antes do primeiro ponto) + parceiro + resto
# Exemplos aceitos:
#   "1234. HRCA-UNI - ..."
#   "1937-1944. DS BELINE - ..."
#   "80-20. HRCA-ARIEL - ..."
LEADING_RE = re.compile(
    r"^\s*(?P<num>[^.]+?)\s*\.\s*(?P<parceiro>.+?)\s+-\s+(?P<resto>.+?)\s*$",
    flags=re.IGNORECASE,
)

# Separadores de polo ativo/passivo (variações)
X_SPLIT_RE = re.compile(
    r"\s+(?:x|X|vs|VS|v\.|V\.|v|V|contra|CONTRA)\s+",
    flags=re.IGNORECASE,
)

MULTISPACE_RE = re.compile(r"\s+")

EXPECTED_HEADERS_PADRAO = [
    "Nº PARCEIRO", "PARCEIRO", "CLIENTE", "CPF", "RÉU", "ESTADO",
    "CLASSIFICAÇÃO", "OBSERVAÇÃO", "RESPONSÁVEL", "NOME DA PASTA",
]

EXPECTED_HEADERS_ROBO = [
    "Pasta",
    "Parceiro",
    "Autor",
    "UF",
]


def detectar_template(col_map: Dict[str, int]) -> str:
    headers = set(col_map.keys())  # já vem normalizado/upper
    # Template "Robô" costuma ter Pasta/Autor/UF
    if "PASTA" in headers and "AUTOR" in headers and "UF" in headers:
        return "ROBO"
    return "PADRAO"


def base_col_para_insercao(col_map: Dict[str, int]) -> int:
    for h in ("Nº PARCEIRO", "PASTA", "NOME DA PASTA"):
        col = col_map.get(h.upper())
        if col is not None:
            return col
    return 1


def norm(s: str) -> str:
    s = (s or "").strip()
    s = MULTISPACE_RE.sub(" ", s)
    return s


# -----------------------------
# MERGE: normalização / lookup
# -----------------------------
def norm_cliente_key(s: str) -> str:
    """
    Normaliza CLIENTE para comparação:
      - trim
      - colapsa espaços
      - upper
      - remove pontuação básica (vira espaço)
    """
    s = norm(s).upper()
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    s = MULTISPACE_RE.sub(" ", s).strip()
    return s


def garantir_coluna(ws: Worksheet, col_map: Dict[str, int], header: str) -> int:
    h = norm(header).upper()
    if h in col_map:
        return col_map[h]

    new_col = ws.max_column + 1
    ws.cell(1, new_col).value = header
    ws.cell(1, new_col).font = Font(bold=True)
    col_map[h] = new_col
    return new_col


def construir_lookup_base(base_path: Path) -> Tuple[Dict[str, Tuple[str, str]], List[str]]:
    """
    Lê a planilha base e cria um índice:
      chave = CLIENTE normalizado
      valor = (ID, CPF)
    """
    avisos: List[str] = []
    wb = openpyxl.load_workbook(base_path, data_only=True)
    ws = wb.active

    col_map = mapear_colunas_por_header(ws)

    def col(*names: str) -> Optional[int]:
        for n in names:
            c = col_map.get(norm(n).upper())
            if c is not None:
                return c
        return None

    c_cliente = col("CLIENTE", "NOME", "AUTOR")
    c_id = col("ID", "ID CLIENTE", "COD", "CÓDIGO", "CODIGO")
    c_cpf = col("CPF", "CPF/CNPJ", "CPF CNPJ", "DOCUMENTO", "DOC")

    if c_cliente is None:
        avisos.append("Planilha base: não encontrei a coluna CLIENTE (ou variações).")
        return {}, avisos

    if c_id is None:
        avisos.append("Planilha base: não encontrei a coluna ID (ou variações).")
    if c_cpf is None:
        avisos.append("Planilha base: não encontrei a coluna CPF (ou variações).")

    lookup: Dict[str, Tuple[str, str]] = {}

    for r in range(2, ws.max_row + 1):
        v_cliente = ws.cell(r, c_cliente).value
        if v_cliente in (None, ""):
            continue

        key = norm_cliente_key(str(v_cliente))
        if not key:
            continue

        v_id = "" if c_id is None else ws.cell(r, c_id).value
        v_cpf = "" if c_cpf is None else ws.cell(r, c_cpf).value

        id_str = norm("" if v_id is None else str(v_id))
        cpf_str = norm("" if v_cpf is None else str(v_cpf))

        if key in lookup:
            if lookup[key] != (id_str, cpf_str):
                avisos.append(f"Base: CLIENTE duplicado com valores diferentes: '{v_cliente}'")
            continue

        lookup[key] = (id_str, cpf_str)

    return lookup, avisos


# -----------------------------
# Parser do nome da pasta
# -----------------------------
def parse_nome_pasta(nome: str) -> Registro:
    """
    Preenche apenas:
      Nº PARCEIRO, PARCEIRO, CLIENTE, RÉU, ESTADO, NOME DA PASTA
    NÃO preenche:
      CLASSIFICAÇÃO, OBSERVAÇÃO, RESPONSÁVEL
    """
    reg = Registro(nome_pasta=nome)
    nome_norm = norm(nome)

    m = LEADING_RE.match(nome_norm)
    if not m:
        reg.warnings.append(
            "Não identifiquei padrão inicial '<id>.<parceiro> - ...' (com ' - ')."
        )
        return reg

    reg.numero_parceiro = norm(m.group("num"))
    reg.parceiro = norm(m.group("parceiro"))  # preserva hífen interno
    resto = norm(m.group("resto"))

    partes = [norm(p) for p in DASH_SEP_RE.split(resto) if norm(p)]
    reg.partes_brutas = partes[:]

    idx_bloco_polos: Optional[int] = None
    for i, p in enumerate(partes):
        if X_SPLIT_RE.search(p):
            idx_bloco_polos = i
            break

    if idx_bloco_polos is not None:
        polos = X_SPLIT_RE.split(partes[idx_bloco_polos], maxsplit=1)
        if len(polos) == 2:
            reg.cliente = norm(polos[0])
            reg.reu = norm(polos[1])
        else:
            reg.warnings.append(
                "Encontrei indicativo de 'X/vs', mas não consegui separar cliente/réu."
            )
            reg.cliente = partes[idx_bloco_polos]
    else:
        reg.warnings.append(
            "Não encontrei separador 'X/vs/contra'; não consegui separar CLIENTE e RÉU."
        )

    for p in partes:
        up = p.upper()
        if UF_RE.match(up):
            reg.estado = up
            break

    if not reg.estado:
        reg.warnings.append("UF/Estado não identificado (esperado 2 letras, ex.: PR).")

    # Garantia: não preencher manualmente
    reg.classificacao = ""
    reg.observacao = ""
    reg.responsavel = ""

    return reg


# -----------------------------
# Validação
# -----------------------------
def validar(reg: Registro) -> List[str]:
    erros: List[str] = []

    if not reg.numero_parceiro:
        erros.append("Faltou Nº PARCEIRO")
    if not reg.parceiro:
        erros.append("Faltou PARCEIRO")
    if not reg.cliente:
        erros.append("Faltou CLIENTE (ou não foi possível separar CLIENTE x RÉU)")
    if not reg.reu:
        erros.append("Faltou RÉU (ou não foi possível separar CLIENTE x RÉU)")
    if not reg.estado:
        erros.append("Faltou ESTADO/UF (2 letras)")

    for w in reg.warnings:
        erros.append(f"Aviso: {w}")

    return erros


def validar_robo(reg: Registro) -> List[str]:
    erros: List[str] = []

    # Para o Robô, validamos só o que será preenchido:
    if not reg.numero_parceiro:
        erros.append("Faltou Pasta/ID (não identifiquei o texto antes do primeiro '.')")
    if not reg.parceiro:
        erros.append("Faltou Parceiro")
    if not reg.cliente:
        erros.append("Faltou Autor (não foi possível identificar o polo ativo)")
    if not reg.estado:
        erros.append("Faltou UF (2 letras)")

    for w in getattr(reg, "warnings", []):
        erros.append(f"Aviso: {w}")

    return erros


# -----------------------------
# Excel helpers
# -----------------------------
def mapear_colunas_por_header(ws: Worksheet) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if isinstance(v, str):
            col_map[norm(v).upper()] = col
    return col_map


def primeira_linha_vazia(ws: Worksheet, col: int, start_row: int = 2) -> int:
    r = start_row
    while True:
        if ws.cell(r, col).value in (None, ""):
            return r
        r += 1


def garantir_aba_validacao(wb: openpyxl.Workbook) -> Worksheet:
    if "VALIDAÇÃO" in wb.sheetnames:
        return wb["VALIDAÇÃO"]

    ws = wb.create_sheet("VALIDAÇÃO")
    ws.append(["NOME DA PASTA", "ERROS/AVISOS", "PARTES INTERPRETADAS"])
    for c in range(1, 4):
        ws.cell(1, c).font = Font(bold=True)
    return ws


def aplicar_marcacao_linha(ws: Worksheet, row: int, is_erro: bool) -> None:
    if not is_erro:
        return
    fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        ws.cell(row, col).fill = fill


def escrever_registro(ws: Worksheet, col_map: Dict[str, int], row: int, reg: Registro) -> None:
    def put(header: str, value: str) -> None:
        col = col_map.get(header.upper())
        if col is None:
            return
        ws.cell(row, col).value = value

    # MERGE (base)
    put("id_Negocio", reg.id_base)
    put("CPF", reg.cpf)

    # Padrão
    put("Nº PARCEIRO", reg.numero_parceiro)
    put("PARCEIRO", reg.parceiro)
    put("CLIENTE", reg.cliente)
    put("RÉU", reg.reu)
    put("ESTADO", reg.estado)

    # Template Robô
    put("PASTA", reg.numero_parceiro)
    put("AUTOR", reg.cliente)
    put("UF", reg.estado)

    # Manual
    put("CLASSIFICAÇÃO", "")
    put("OBSERVAÇÃO", "")
    put("RESPONSÁVEL", "")

    put("NOME DA PASTA", reg.nome_pasta)


def listar_subpastas(diretorio: Path) -> List[Path]:
    if not diretorio.is_dir():
        raise ValueError(f"Diretório inválido: {diretorio}")
    return sorted([p for p in diretorio.iterdir() if p.is_dir()])


# -----------------------------
# Função principal (agora com merge)
# -----------------------------
def preencher_excel_parceiro(
    template_path: Path,
    pasta_raiz: Path,
    output_path: Path,
    base_xlsx_path: Optional[Path] = None,  # NOVO
) -> tuple[int, int]:
    """
    Motor do preenchimento.
    Retorna: (linhas_ok, linhas_com_problemas).
    """
    subpastas = listar_subpastas(pasta_raiz)
    registros = [parse_nome_pasta(p.name) for p in subpastas]

    # 1) Carrega lookup da planilha base (se informada)
    lookup_base: Dict[str, Tuple[str, str]] = {}
    avisos_base: List[str] = []
    if base_xlsx_path is not None:
        if not base_xlsx_path.exists():
            raise FileNotFoundError(f"Planilha base não encontrada: {base_xlsx_path}")
        lookup_base, avisos_base = construir_lookup_base(base_xlsx_path)

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    col_map = mapear_colunas_por_header(ws)

    kind = detectar_template(col_map)
    expected = EXPECTED_HEADERS_ROBO if kind == "ROBO" else EXPECTED_HEADERS_PADRAO
    faltantes = [h for h in expected if h.upper() not in col_map]

    # 2) Garante as colunas de saída do merge
    # garantir_coluna(ws, col_map, "ID")
    garantir_coluna(ws, col_map, "CPF")

    base_col = base_col_para_insercao(col_map)
    row = primeira_linha_vazia(ws, col=base_col, start_row=2)

    ws_val = garantir_aba_validacao(wb)

    # registra avisos da base na aba VALIDACAO
    for av in avisos_base:
        ws_val.append(["(PLANILHA BASE)", f"Aviso: {av}", ""])

    ok = 0
    problemas = 0

    for reg in registros:
        # 3) Merge: CLIENTE (pasta) -> (ID, CPF) (base)
        if lookup_base:
            key = norm_cliente_key(reg.cliente)
            if key in lookup_base:
                reg.id_base, reg.cpf = lookup_base[key]
            else:
                reg.warnings.append("Não encontrei CLIENTE na planilha base para trazer ID/CPF.")

        erros = validar_robo(reg) if kind == "ROBO" else validar(reg)
        tem_problema = len(erros) > 0

        escrever_registro(ws, col_map, row, reg)
        aplicar_marcacao_linha(ws, row, is_erro=tem_problema)

        if tem_problema:
            problemas += 1
            ws_val.append([reg.nome_pasta, "; ".join(erros), " | ".join(reg.partes_brutas)])
        else:
            ok += 1

        row += 1

    if faltantes:
        ws_val.insert_rows(2)
        ws_val.cell(2, 1).value = "AVISO TEMPLATE"
        ws_val.cell(2, 2).value = (
            f"Colunas não encontradas no template (linha 1): {', '.join(faltantes)}"
        )
        ws_val.cell(2, 3).value = "A escrita nessas colunas foi ignorada."
        for c in range(1, 4):
            ws_val.cell(2, c).font = Font(bold=True)

    wb.save(output_path)
    return ok, problemas